import io
import calendar
from datetime import datetime, timezone
from typing import Optional, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database import get_db, get_vn_now, VN_TZ
from config import settings

def to_vn_dt(dt):
    """Chuyển đổi datetime về múi giờ Việt Nam (UTC+7) chuẩn xác."""
    if dt is None:
        return None
    if getattr(dt, 'tzinfo', None) is not None:
        return dt.astimezone(VN_TZ)
    # Nếu là naive datetime (ví dụ bản ghi cũ lưu theo UTC):
    return dt.replace(tzinfo=timezone.utc).astimezone(VN_TZ)

class ReportService:
    @staticmethod
    async def get_monthly_summary(year: int, month: int):
        """Tính toán số liệu báo cáo Uptime/Downtime trong tháng theo múi giờ Việt Nam (UTC+7)."""
        db = get_db()
        _, num_days = calendar.monthrange(year, month)
        start_vn = datetime(year, month, 1, 0, 0, 0, tzinfo=VN_TZ)
        end_vn = datetime(year, month, num_days, 23, 59, 59, tzinfo=VN_TZ)
        start_utc_naive = start_vn.astimezone(timezone.utc).replace(tzinfo=None)
        end_utc_naive = end_vn.astimezone(timezone.utc).replace(tzinfo=None)
        total_month_seconds = num_days * 86400

        # Lấy tất cả thiết bị và kênh
        devices = await db.devices.find().to_list(100)
        channels = await db.channels.find().to_list(500)

        # Lấy tất cả các sự cố trong tháng (hỗ trợ cả aware VN_TZ lẫn naive UTC)
        events_cursor = db.events.find({
            "$or": [
                {"timestamp": {"$gte": start_vn, "$lte": end_vn}},
                {"timestamp": {"$gte": start_utc_naive, "$lte": end_utc_naive}}
            ],
            "event": {"$in": ["offline", "video_loss"]}
        })
        events = await events_cursor.to_list(2000)
        now_vn = get_vn_now()

        # Bỏ qua các sự cố gián đoạn quá ngắn dưới ngưỡng (ví dụ < 30 phút)
        min_sec = getattr(settings, "min_incident_seconds", 1800)
        valid_events = []
        for e in events:
            dur = e.get("duration_seconds")
            if dur is not None:
                if dur >= min_sec:
                    valid_events.append(e)
            else:
                ev_time = to_vn_dt(e.get("timestamp")) or start_vn
                ongoing_sec = max(0, int((now_vn - ev_time).total_seconds()))
                if ongoing_sec >= min_sec:
                    valid_events.append(e)
        events = valid_events

        # Thống kê cho từng đầu thu
        device_reports = []
        for dev in devices:
            dev_id = str(dev["_id"])
            dev_events = [
                e for e in events 
                if (e.get("target_type") == "device" and str(e.get("target_id")) == dev_id)
                or (str(e.get("device_id")) == dev_id)
            ]
            downtime_sec = 0
            for ev in dev_events:
                dur = ev.get("duration_seconds")
                if dur is None:
                    ev_time = to_vn_dt(ev.get("timestamp")) or start_vn
                    calc_end = min(now_vn, end_vn)
                    dur = max(0, int((calc_end - ev_time).total_seconds()))
                downtime_sec += dur

            downtime_sec = min(downtime_sec, total_month_seconds)
            uptime_pct = round(((total_month_seconds - downtime_sec) / total_month_seconds) * 100, 2)

            device_reports.append({
                "id": dev_id,
                "name": dev.get("name", "NVR"),
                "ip": dev.get("ip"),
                "location": dev.get("location", "LAN"),
                "channel_count": dev.get("channel_count", 0),
                "incident_count": len(dev_events),
                "downtime_seconds": downtime_sec,
                "downtime_minutes": round(downtime_sec / 60, 1),
                "downtime_hours": round(downtime_sec / 3600, 2),
                "uptime_percent": uptime_pct,
                "status": dev.get("status", "unknown")
            })

        # Thống kê cho từng kênh camera
        channel_reports = []
        for ch in channels:
            ch_id = str(ch["_id"])
            ch_dev_id = str(ch.get("device_id", ""))
            ch_num = ch.get("channel_no")
            ch_events = [
                e for e in events 
                if (e.get("target_type") == "channel" and (str(e.get("target_id")) == ch_id or (str(e.get("device_id")) == ch_dev_id and e.get("channel_no") == ch_num)))
                or (e.get("target_type") == "device" and (str(e.get("target_id")) == ch_dev_id or str(e.get("device_id")) == ch_dev_id))
            ]
            downtime_sec = 0
            for ev in ch_events:
                dur = ev.get("duration_seconds")
                if dur is None:
                    ev_time = to_vn_dt(ev.get("timestamp")) or start_vn
                    calc_end = min(now_vn, end_vn)
                    dur = max(0, int((calc_end - ev_time).total_seconds()))
                downtime_sec += dur

            downtime_sec = min(downtime_sec, total_month_seconds)
            uptime_pct = round(((total_month_seconds - downtime_sec) / total_month_seconds) * 100, 2)

            channel_reports.append({
                "id": ch_id,
                "device_id": ch.get("device_id"),
                "device_name": ch.get("device_name", ""),
                "channel_no": ch.get("channel_no"),
                "name": ch.get("name", f"Camera {ch.get('channel_no')}"),
                "incident_count": len(ch_events),
                "downtime_seconds": downtime_sec,
                "downtime_minutes": round(downtime_sec / 60, 1),
                "downtime_hours": round(downtime_sec / 3600, 2),
                "uptime_percent": uptime_pct,
                "status": ch.get("status", "online")
            })

        # Tính tổng quan toàn hệ thống
        all_uptimes = [d["uptime_percent"] for d in device_reports] + [c["uptime_percent"] for c in channel_reports]
        avg_system_uptime = round(sum(all_uptimes) / len(all_uptimes), 2) if all_uptimes else 100.0
        total_incidents = len(events)

        return {
            "year": year,
            "month": month,
            "total_days": num_days,
            "total_devices": len(devices),
            "total_channels": len(channels),
            "avg_uptime_percent": avg_system_uptime,
            "total_incidents": total_incidents,
            "devices": device_reports,
            "channels": channel_reports,
            "recent_events": [
                {
                    "id": str(e["_id"]),
                    "target_name": e.get("target_name"),
                    "event": e.get("event"),
                    "timestamp": to_vn_dt(e.get("timestamp")).isoformat() if e.get("timestamp") else None,
                    "resolved_at": to_vn_dt(e.get("resolved_at")).isoformat() if e.get("resolved_at") else None,
                    "duration_seconds": e.get("duration_seconds"),
                    "note": e.get("note")
                }
                for e in events[:50]
            ]
        }

    @staticmethod
    async def export_excel_report(year: int, month: int) -> io.BytesIO:
        """Xuất file Excel báo cáo tháng định dạng chuyên nghiệp."""
        data = await ReportService.get_monthly_summary(year, month)

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Báo cáo Tổng hợp"

        # Styles
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
        sub_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Blue
        title_font = Font(name="Arial", size=15, bold=True, color="1E3A8A")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Arial", size=10, bold=True)
        normal_font = Font(name="Arial", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

        # Tiêu đề báo cáo
        ws1.merge_cells("A1:G1")
        ws1["A1"] = f"BÁO CÁO HOẠT ĐỘNG & TỶ LỆ SẴN SÀNG (SLA) HỆ THỐNG CAMERA"
        ws1["A1"].font = title_font
        ws1["A1"].alignment = center_align

        ws1.merge_cells("A2:G2")
        ws1["A2"] = f"Tháng: {month:02d}/{year} | Tổng số ngày: {data['total_days']} ngày | Uptime toàn hệ thống: {data['avg_uptime_percent']}%"
        ws1["A2"].font = Font(name="Arial", size=11, italic=True, color="4B5563")
        ws1["A2"].alignment = center_align

        ws1.row_dimensions[1].height = 30
        ws1.row_dimensions[2].height = 20

        # PHẦN 1: BẢNG ĐẦU THU (NVR)
        row = 4
        ws1.cell(row=row, column=1, value="I. DANH SÁCH ĐẦU THU (NVR/DVR)").font = bold_font
        row += 1

        nvr_headers = ["STT", "Tên Đầu Thu", "Địa chỉ IP", "Vị trí / Mạng", "Số Kênh", "Số Lần Sự Cố", "Thời Gian Gián Đoạn (giờ)", "Tỷ Lệ Uptime (%)", "Đánh Giá"]
        for col_idx, h in enumerate(nvr_headers, start=1):
            c = ws1.cell(row=row, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border
        ws1.row_dimensions[row].height = 24

        for idx, dev in enumerate(data["devices"], start=1):
            row += 1
            ws1.cell(row=row, column=1, value=idx).alignment = center_align
            ws1.cell(row=row, column=2, value=dev["name"]).alignment = left_align
            ws1.cell(row=row, column=3, value=dev["ip"]).alignment = center_align
            ws1.cell(row=row, column=4, value=dev["location"]).alignment = center_align
            ws1.cell(row=row, column=5, value=dev["channel_count"]).alignment = center_align
            ws1.cell(row=row, column=6, value=dev["incident_count"]).alignment = center_align
            ws1.cell(row=row, column=7, value=dev["downtime_hours"]).alignment = right_align
            
            uptime_cell = ws1.cell(row=row, column=8, value=f"{dev['uptime_percent']}%")
            uptime_cell.alignment = center_align
            uptime_cell.font = bold_font

            eval_text = "Đạt SLA (>=99%)" if dev["uptime_percent"] >= 99.0 else "Cần bảo trì (<99%)"
            ws1.cell(row=row, column=9, value=eval_text).alignment = center_align

            for c_idx in range(1, 10):
                ws1.cell(row=row, column=c_idx).border = thin_border
                ws1.cell(row=row, column=c_idx).font = normal_font

        # PHẦN 2: BẢNG CHI TIẾT TỪNG KÊNH CAMERA
        row += 2
        ws1.cell(row=row, column=1, value="II. DANH SÁCH CHI TIẾT TỪNG MẮT CAMERA").font = bold_font
        row += 1

        cam_headers = ["STT", "Tên Camera", "Thuộc Đầu Thu", "Kênh Số", "Số Lần Mất Tín Hiệu", "Thời Gian Mất Tín Hiệu (giờ)", "Tỷ Lệ Uptime (%)", "Trạng Thái Cuối"]
        for col_idx, h in enumerate(cam_headers, start=1):
            c = ws1.cell(row=row, column=col_idx, value=h)
            c.font = header_font
            c.fill = sub_fill
            c.alignment = center_align
            c.border = thin_border
        ws1.row_dimensions[row].height = 24

        for idx, cam in enumerate(data["channels"], start=1):
            row += 1
            ws1.cell(row=row, column=1, value=idx).alignment = center_align
            ws1.cell(row=row, column=2, value=cam["name"]).alignment = left_align
            ws1.cell(row=row, column=3, value=cam["device_name"]).alignment = left_align
            ws1.cell(row=row, column=4, value=f"Kênh {cam['channel_no']}").alignment = center_align
            ws1.cell(row=row, column=5, value=cam["incident_count"]).alignment = center_align
            ws1.cell(row=row, column=6, value=cam["downtime_hours"]).alignment = right_align
            
            c_uptime = ws1.cell(row=row, column=7, value=f"{cam['uptime_percent']}%")
            c_uptime.alignment = center_align
            c_uptime.font = bold_font

            st_text = "Bình thường" if cam["status"] == "online" else "Mất tín hiệu"
            ws1.cell(row=row, column=8, value=st_text).alignment = center_align

            for c_idx in range(1, 9):
                ws1.cell(row=row, column=c_idx).border = thin_border
                ws1.cell(row=row, column=c_idx).font = normal_font

        # Auto-adjust column widths
        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # SHEET 2: NHẬT KÝ SỰ CỐ CHI TIẾT
        ws2 = wb.create_sheet(title="Nhật Ký Sự Cố")
        ws2.merge_cells("A1:F1")
        ws2["A1"] = f"NHẬT KÝ CHI TIẾT SỰ CỐ MẤT TÍN HIỆU - THÁNG {month:02d}/{year}"
        ws2["A1"].font = title_font
        ws2["A1"].alignment = center_align
        ws2.row_dimensions[1].height = 28

        log_headers = ["STT", "Thời Điểm Bắt Đầu", "Thời Điểm Phục Hồi", "Thời Lượng Gián Đoạn", "Đối Tượng", "Chi Tiết Sự Cố"]
        for col_idx, h in enumerate(log_headers, start=1):
            c = ws2.cell(row=3, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border
        ws2.row_dimensions[3].height = 24

        if not data["recent_events"]:
            ws2.merge_cells("A4:F4")
            c_empty = ws2.cell(row=4, column=1, value="Hệ thống camera hoạt động liên tục 24/7 trong tháng, không phát sinh sự cố nào.")
            c_empty.alignment = center_align
            c_empty.font = Font(name="Arial", size=10, italic=True, color="059669")
            ws2.row_dimensions[4].height = 24
            for c_idx in range(1, 7):
                ws2.cell(row=4, column=c_idx).border = thin_border
        else:
            for idx, ev in enumerate(data["recent_events"], start=1):
                r = 3 + idx
                ws2.cell(row=r, column=1, value=idx).alignment = center_align

                ts_str = "-"
                if ev.get("timestamp"):
                    try:
                        dt_val = datetime.fromisoformat(ev["timestamp"])
                        ts_str = dt_val.strftime("%d/%m/%Y %H:%M:%S")
                    except Exception:
                        ts_str = ev["timestamp"][:19].replace("T", " ")

                res_str = "Đang gián đoạn..."
                if ev.get("resolved_at"):
                    try:
                        dt_res = datetime.fromisoformat(ev["resolved_at"])
                        res_str = dt_res.strftime("%d/%m/%Y %H:%M:%S")
                    except Exception:
                        res_str = ev["resolved_at"][:19].replace("T", " ")

                ws2.cell(row=r, column=2, value=ts_str).alignment = center_align
                ws2.cell(row=r, column=3, value=res_str).alignment = center_align
                
                dur = ev.get("duration_seconds")
                dur_str = f"{round(dur/60, 1)} phút" if dur is not None else "Đang gián đoạn"
                ws2.cell(row=r, column=4, value=dur_str).alignment = center_align
                ws2.cell(row=r, column=5, value=ev.get("target_name", "")).alignment = left_align
                ws2.cell(row=r, column=6, value=ev.get("note", "")).alignment = left_align

                for c_idx in range(1, 7):
                    ws2.cell(row=r, column=c_idx).border = thin_border
                    ws2.cell(row=r, column=c_idx).font = normal_font

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 14)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    async def export_haiquan_daily_excel(year: int, month: int, reporter: str = "") -> io.BytesIO:
        """
        Xuất file Excel theo đúng chuẩn biểu mẫu:
        'KIỂM TRA HOẠT ĐỘNG CỦA CAMERA HẢI QUAN'
        Gồm 31 cột ngày (1-31), S/N Đầu thu, S/N Camera, Khu vực, Tên camera và bảng Ghi chú chân trang.
        """
        db = get_db()
        _, num_days = calendar.monthrange(year, month)
        start_month = datetime(year, month, 1, 0, 0, 0, tzinfo=VN_TZ)
        end_month = datetime(year, month, num_days, 23, 59, 59, tzinfo=VN_TZ)
        start_utc_naive = start_month.astimezone(timezone.utc).replace(tzinfo=None)
        end_utc_naive = end_month.astimezone(timezone.utc).replace(tzinfo=None)

        devices = await db.devices.find().sort("name", 1).to_list(100)
        channels = await db.channels.find().sort([("device_id", 1), ("channel_no", 1)]).to_list(500)

        # Lấy tất cả sự cố trong tháng (hỗ trợ cả aware VN_TZ lẫn naive UTC)
        events_cursor = db.events.find({
            "$or": [
                {"timestamp": {"$gte": start_month, "$lte": end_month}},
                {"timestamp": {"$gte": start_utc_naive, "$lte": end_utc_naive}}
            ],
            "event": {"$in": ["offline", "video_loss"]}
        }).sort("timestamp", 1)
        events = await events_cursor.to_list(2000)
        now_vn = get_vn_now()

        # Bỏ qua sự cố gián đoạn ngắn dưới ngưỡng (< 30 phút)
        min_sec = getattr(settings, "min_incident_seconds", 1800)
        valid_events = []
        for e in events:
            dur = e.get("duration_seconds")
            if dur is not None:
                if dur >= min_sec:
                    valid_events.append(e)
            else:
                ev_time = to_vn_dt(e.get("timestamp")) or start_month
                ongoing_sec = max(0, int((now_vn - ev_time).total_seconds()))
                if ongoing_sec >= min_sec:
                    valid_events.append(e)
        events = valid_events

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"HaiQuan_T{month:02d}_{year}"

        # Cài đặt trang in ngang A4
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # Fonts & Styles
        font_title = Font(name="Times New Roman", size=14, bold=True)
        font_header = Font(name="Times New Roman", size=9, bold=True)
        font_bold = Font(name="Times New Roman", size=9, bold=True)
        font_regular = Font(name="Times New Roman", size=8.5)
        font_note_bold = Font(name="Times New Roman", size=9, bold=True)
        font_note_item = Font(name="Times New Roman", size=8.5)

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        thin = Side(style='thin', color='000000')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 1. TIÊU ĐỀ BÁO CÁO (Ô A1:D2)
        ws.merge_cells("A1:D2")
        title_cell = ws["A1"]
        title_cell.value = "KIỂM TRA HOẠT ĐỘNG CỦA CAMERA HẢI QUAN"
        title_cell.font = font_title
        title_cell.alignment = align_center

        # 2. KHỐI THÔNG TIN THÁNG & NGƯỜI THỰC HIỆN (Góc trên bên phải)
        # Cột AE (31) đến AI (35)
        ws.cell(row=1, column=31, value="Tháng / năm").font = font_bold
        ws.cell(row=1, column=31).alignment = align_center
        ws.cell(row=1, column=31).border = border_all

        ws.cell(row=1, column=32, value="Người thực hiện").font = font_bold
        ws.cell(row=1, column=32).alignment = align_center
        ws.cell(row=1, column=32).border = border_all
        ws.merge_cells(start_row=1, start_column=32, end_row=1, end_column=35)

        ws.cell(row=2, column=31, value=f"{month:02d}/{year}").font = font_regular
        ws.cell(row=2, column=31).alignment = align_center
        ws.cell(row=2, column=31).border = border_all

        ws.cell(row=2, column=32, value=reporter or "").font = font_regular
        ws.cell(row=2, column=32).alignment = align_center
        ws.cell(row=2, column=32).border = border_all
        ws.merge_cells(start_row=2, start_column=32, end_row=2, end_column=35)

        # 3. TIÊU ĐỀ CÁC CỘT DỮ LIỆU (Row 4)
        headers = ["S/N đầu thu", "S/N Camera", "Khu vực", "Tên camera"] + [str(d) for d in range(1, 32)]
        ws.row_dimensions[4].height = 26

        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=4, column=col_idx, value=h)
            c.font = font_header
            c.alignment = align_center
            c.border = border_all

        # 4. ĐỔ DỮ LIỆU TỪNG ĐẦU THU VÀ CAMERA
        row = 4
        # Đánh số thứ tự duy nhất (1, 2, 3...) cho từng sự cố trong tháng
        event_num_map = {}
        incident_notes = []
        for idx, ev in enumerate(events, start=1):
            ev_id_str = str(ev["_id"])
            event_num_map[ev_id_str] = idx

            t_start = to_vn_dt(ev.get("timestamp"))
            t_end = to_vn_dt(ev.get("resolved_at"))
            t_start_str = t_start.strftime("%H:%M ngày %d/%m/%Y") if t_start else "N/A"
            t_end_str = t_end.strftime("%H:%M ngày %d/%m/%Y") if t_end else "Đang khắc phục"
            target_desc = ev.get("target_name") or f"Kênh {ev.get('channel_no', '')}"
            note_text = ev.get("note") or f"Mất tín hiệu camera ({target_desc})"
            incident_notes.append(f'"{idx}" - "{t_start_str}" - "{t_end_str}" - "{note_text}"')

        for dev in devices:
            dev_id = str(dev["_id"])
            dev_name = dev.get("name", "NVR")
            dev_sn = dev.get("serial_no") or dev.get("serial") or "6L0651EPAZC3D7B"
            dev_channels = [c for c in channels if c.get("device_id") == dev_id and c.get("enabled", True) and c.get("status") != "unconnected"]

            if not dev_channels:
                continue

            start_dev_row = row + 1

            for ch in dev_channels:
                row += 1
                ws.row_dimensions[row].height = 19
                ch_id = str(ch["_id"])
                ch_name = ch.get("name", f"Camera {ch.get('channel_no')}")
                cam_sn = ch.get("serial_no") or ch.get("serial") or f"7D06605PAG{ch.get('channel_no'):02d}"
                zone = ch.get("zone") or dev.get("location") or "TOWA HCM 1"

                # S/N Camera
                c_cam = ws.cell(row=row, column=2, value=cam_sn)
                c_cam.font = font_regular
                c_cam.alignment = align_center
                c_cam.border = border_all

                # Khu vực
                c_zone = ws.cell(row=row, column=3, value=zone)
                c_zone.font = font_regular
                c_zone.alignment = align_center
                c_zone.border = border_all

                # Tên camera
                c_name = ws.cell(row=row, column=4, value=ch_name)
                c_name.font = font_regular
                c_name.alignment = align_left
                c_name.border = border_all

                # Check sự cố từng ngày từ 1 đến 31
                ch_num = ch.get("channel_no")
                ch_events = [
                    e for e in events 
                    if (e.get("target_type") == "channel" and (str(e.get("target_id")) == ch_id or (str(e.get("device_id")) == dev_id and e.get("channel_no") == ch_num)))
                    or (e.get("target_type") == "device" and (str(e.get("target_id")) == dev_id or str(e.get("device_id")) == dev_id))
                ]

                for day in range(1, 32):
                    col_day = 4 + day
                    c_day = ws.cell(row=row, column=col_day)
                    c_day.border = border_all
                    c_day.alignment = align_center

                    if day > num_days:
                        # Những ngày không tồn tại trong tháng (ví dụ ngày 31 tháng 4)
                        c_day.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                        continue

                    day_start = datetime(year, month, day, 0, 0, 0, tzinfo=VN_TZ)
                    day_end = datetime(year, month, day, 23, 59, 59, tzinfo=VN_TZ)

                    matching_events = []
                    for ev in ch_events:
                        ev_start = to_vn_dt(ev.get("timestamp"))
                        ev_end = to_vn_dt(ev.get("resolved_at")) or end_month
                        if ev_start and ev_start <= day_end and ev_end >= day_start:
                            matching_events.append(ev)

                    if matching_events:
                        cell_labels = [str(event_num_map.get(str(ev["_id"]), "?")) for ev in matching_events]
                        c_day.value = ",".join(cell_labels)
                        c_day.font = Font(name="Times New Roman", size=8, bold=True, color="DC2626")
                    else:
                        # Hoạt động đủ 24/7 -> để trống theo chuẩn form Hải Quan
                        c_day.value = ""

            end_dev_row = row
            # Merge cột S/N đầu thu cho toàn bộ các dòng của cùng đầu thu đó
            ws.merge_cells(start_row=start_dev_row, start_column=1, end_row=end_dev_row, end_column=1)
            c_dev = ws.cell(row=start_dev_row, column=1)
            c_dev.value = f"{dev_sn}\n( {dev_name} )"
            c_dev.font = font_bold
            c_dev.alignment = align_center

            for r_idx in range(start_dev_row, end_dev_row + 1):
                ws.cell(row=r_idx, column=1).border = border_all

        # 5. KHỐI GHI CHÚ CHÂN TRANG (Footnotes)
        row += 2
        ws.cell(row=row, column=1, value='Ghi chú: (Đánh số vào các ô tương ứng khi không đủ dữ liệu 24/7 và ghi theo cú pháp). Nội dung ghi chú: "Số" - "Giờ, phút bắt đầu" - "Giờ, phút kết thúc" - "Lý do". Trường hợp mất dữ liệu theo nhóm đầu thu thì thay "Số" bằng "Tên đầu thu"').font = font_note_bold

        if incident_notes:
            for item in incident_notes:
                row += 1
                ws.cell(row=row, column=1, value=item).font = font_note_item
        else:
            row += 1
            ws.cell(row=row, column=1, value="Hệ thống camera hoạt động liên tục 24/7 trong tháng, không phát sinh sự cố gián đoạn dữ liệu.").font = font_regular

        # Độ rộng các cột
        ws.column_dimensions["A"].width = 19
        ws.column_dimensions["B"].width = 19
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 23
        for d in range(1, 32):
            col_letter = openpyxl.utils.get_column_letter(4 + d)
            ws.column_dimensions[col_letter].width = 3.6

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    async def get_data_retention_stats(year: Optional[int] = None):
        """Lấy thống kê số lượng sự cố theo từng tháng (bao gồm cả các tháng 0 sự cố) để người dùng tải báo cáo và quản lý chốt sổ."""
        db = get_db()
        vn_now = get_vn_now()
        current_year = vn_now.year

        if year is not None:
            years_to_process = [year]
        else:
            # Lấy tất cả các năm có trong db.events + năm hiện tại
            pipeline_years = [
                {"$match": {"timestamp": {"$ne": None}}},
                {
                    "$group": {
                        "_id": {
                            "$year": {"date": "$timestamp", "timezone": "+07:00"}
                        }
                    }
                },
                {"$sort": {"_id": -1}}
            ]
            try:
                year_results = await db.events.aggregate(pipeline_years).to_list(50)
                years_in_db = [r["_id"] for r in year_results if r.get("_id") is not None]
            except Exception:
                years_in_db = []

            years_to_process = sorted(list(set(years_in_db + [current_year])), reverse=True)

        # Thống kê số sự cố theo từng (năm, tháng)
        pipeline = [
            {"$match": {"timestamp": {"$ne": None}}},
            {
                "$group": {
                    "_id": {
                        "year": {"$year": {"date": "$timestamp", "timezone": "+07:00"}},
                        "month": {"$month": {"date": "$timestamp", "timezone": "+07:00"}}
                    },
                    "count": {"$sum": 1},
                    "resolved_count": {
                        "$sum": {"$cond": [{"$ne": ["$resolved_at", None]}, 1, 0]}
                    }
                }
            }
        ]
        try:
            results = await db.events.aggregate(pipeline).to_list(500)
            events_by_ym = {
                (r["_id"]["year"], r["_id"]["month"]): r
                for r in results
                if r.get("_id") and r["_id"].get("year") is not None and r["_id"].get("month") is not None
            }
        except Exception:
            events_by_ym = {}

        retention_list = []
        for y in years_to_process:
            for m in range(12, 0, -1):
                rec = events_by_ym.get((y, m))
                total = rec["count"] if rec else 0
                resolved = rec["resolved_count"] if rec else 0
                retention_list.append({
                    "year": y,
                    "month": m,
                    "total_events": total,
                    "resolved_events": resolved
                })

        return retention_list

    @staticmethod
    async def delete_monthly_events(year: int, month: int):
        """Xóa dữ liệu sự cố của một tháng sau khi đã chốt sổ và tải về máy."""
        db = get_db()
        _, num_days = calendar.monthrange(year, month)
        start_vn = datetime(year, month, 1, 0, 0, 0, tzinfo=VN_TZ)
        end_vn = datetime(year, month, num_days, 23, 59, 59, tzinfo=VN_TZ)
        start_utc = start_vn.astimezone(timezone.utc)
        end_utc = end_vn.astimezone(timezone.utc)
        start_utc_naive = start_utc.replace(tzinfo=None)
        end_utc_naive = end_utc.replace(tzinfo=None)

        result = await db.events.delete_many({
            "$or": [
                {"timestamp": {"$gte": start_vn, "$lte": end_vn}},
                {"timestamp": {"$gte": start_utc, "$lte": end_utc}},
                {"timestamp": {"$gte": start_utc_naive, "$lte": end_utc_naive}}
            ]
        })
        return {
            "success": True,
            "deleted_count": result.deleted_count,
            "message": f"Đã xóa hoàn tất {result.deleted_count} bản ghi sự cố của Tháng {month:02d}/{year}."
        }

    @staticmethod
    async def delete_yearly_events(year: int):
        """Xóa toàn bộ dữ liệu sự cố của cả năm."""
        db = get_db()
        start_vn = datetime(year, 1, 1, 0, 0, 0, tzinfo=VN_TZ)
        end_vn = datetime(year, 12, 31, 23, 59, 59, tzinfo=VN_TZ)
        start_utc = start_vn.astimezone(timezone.utc)
        end_utc = end_vn.astimezone(timezone.utc)
        start_utc_naive = start_utc.replace(tzinfo=None)
        end_utc_naive = end_utc.replace(tzinfo=None)

        result = await db.events.delete_many({
            "$or": [
                {"timestamp": {"$gte": start_vn, "$lte": end_vn}},
                {"timestamp": {"$gte": start_utc, "$lte": end_utc}},
                {"timestamp": {"$gte": start_utc_naive, "$lte": end_utc_naive}}
            ]
        })
        return {
            "success": True,
            "deleted_count": result.deleted_count,
            "message": f"Đã xóa hoàn tất {result.deleted_count} bản ghi sự cố của Năm {year}."
        }

